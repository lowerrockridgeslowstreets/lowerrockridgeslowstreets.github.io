"""
Build ADT summaries workbook from Telraam API daily CSVs in API_Data.

For each street:
  1) Embed telraam_*_data.csv into *_pull, duplicate to *_data.
  2) *_daily sums modes from *_data via SUMIFS by calendar date.

Output:
    ``<DATA_DIR>/adt_summaries_v5.xlsx`` (``DATA_DIR`` from ``telraam_paths`` / ``TELRAAM_DATA_DIR``)

Refresh API CSVs (telraam_code pull scripts), then:
    python3 telraam_code/build_adt_summaries_xlsx.py
"""

from __future__ import annotations

import csv
import sys
from datetime import date, datetime
from pathlib import Path

_script_dir = Path(__file__).resolve().parent
if str(_script_dir) not in sys.path:
    sys.path.insert(0, str(_script_dir))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from telraam_paths import get_telraam_data_dir

DATA_DIR = get_telraam_data_dir()
API_DATA_DIR = DATA_DIR / "API_Data"
COLBY_CSV = API_DATA_DIR / "telraam_colby_data.csv"
HILL_CSV = API_DATA_DIR / "telraam_hillegass_9000008271_data.csv"
MARTIN_CSV = API_DATA_DIR / "telraam_martin_data.csv"
OUT = DATA_DIR / "adt_summaries_v5.xlsx"


def embed_api_daily_csv(
    wb: Workbook,
    csv_path: Path,
    pull_name: str,
    data_name: str,
) -> tuple[list[str], dict[str, int]]:
    """Embed API CSV into pull + data sheets; return sorted YYYY-MM-DD keys and header indices."""
    ws = wb.create_sheet(pull_name)
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    if not rows:
        raise SystemExit(f"Empty CSV: {csv_path}")
    header = [h.strip() for h in rows[0]]
    idx = {h: i + 1 for i, h in enumerate(header)}
    need = ("date", "pedestrian", "bike", "car", "heavy", "night")
    for k in need:
        if k not in idx:
            raise SystemExit(f"{csv_path.name} missing column {k!r}; have {list(idx)}")
    date_col_idx = header.index("date")
    numeric_keys = ("pedestrian", "bike", "car", "heavy", "night")

    ws.append(header)
    seen: set[str] = set()
    for raw in rows[1:]:
        if not raw or len(raw) < len(header):
            continue
        pad = raw + [""] * (len(header) - len(raw))
        row = pad[: len(header)]
        ds = str(row[date_col_idx]).strip()
        for nk in numeric_keys:
            i = header.index(nk)
            v = row[i]
            if v == "" or v is None:
                row[i] = 0.0
            else:
                row[i] = float(v)
        row[date_col_idx] = datetime.strptime(ds, "%Y-%m-%d").date()
        ws.append(row)
        seen.add(ds)
    ws_data = wb.copy_worksheet(ws)
    ws_data.title = data_name
    return sorted(seen), idx


def fill_daily_sheet(
    wb: Workbook,
    date_keys: list[str],
    col_idx: dict[str, int],
    daily_name: str,
    data_name: str,
) -> None:
    ws = wb.create_sheet(daily_name)
    date_col = get_column_letter(col_idx["date"])
    car_col = get_column_letter(col_idx["car"])
    heavy_col = get_column_letter(col_idx["heavy"])
    night_col = get_column_letter(col_idx["night"])
    ped_col = get_column_letter(col_idx["pedestrian"])
    bike_col = get_column_letter(col_idx["bike"])

    hdr = [
        "Date",
        "Cars",
        "Large",
        "Night",
        "Ped",
        "Bike",
        "Motorized",
        "All_modes",
        "Weekday_1_to_7",
        "Week_key",
    ]
    for c, h in enumerate(hdr, 1):
        ws.cell(1, c, h)

    mp = data_name
    r = 2
    for d in date_keys:
        ac = f"A{r}"
        ws.cell(r, 1, datetime.strptime(d, "%Y-%m-%d").date())
        ws.cell(
            r,
            2,
            f"=SUMIFS({mp}!${car_col}:${car_col},{mp}!${date_col}:${date_col},{ac})",
        )
        ws.cell(
            r,
            3,
            f"=SUMIFS({mp}!${heavy_col}:${heavy_col},{mp}!${date_col}:${date_col},{ac})",
        )
        ws.cell(
            r,
            4,
            f"=SUMIFS({mp}!${night_col}:${night_col},{mp}!${date_col}:${date_col},{ac})",
        )
        ws.cell(
            r,
            5,
            f"=SUMIFS({mp}!${ped_col}:${ped_col},{mp}!${date_col}:${date_col},{ac})",
        )
        ws.cell(
            r,
            6,
            f"=SUMIFS({mp}!${bike_col}:${bike_col},{mp}!${date_col}:${date_col},{ac})",
        )
        ws.cell(r, 7, f"=B{r}+C{r}+D{r}")
        ws.cell(r, 8, f"=G{r}+E{r}+F{r}")
        ws.cell(r, 9, f"=WEEKDAY(A{r},2)")
        ws.cell(r, 10, f'=TEXT(A{r},"yyyy")&"-W"&TEXT(WEEKNUM(A{r},21),"00")')
        r += 1


def fill_weekly(wb: Workbook, ws_name: str, daily_sheet: str, dates: list[date]) -> None:
    ws = wb.create_sheet(ws_name)
    ws.cell(1, 1, "Week_key")
    ws.cell(1, 2, "Avg_motorized")
    keys: set[str] = set()
    for d in dates:
        iso = d.isocalendar()
        keys.add(f"{iso[0]}-W{iso[1]:02d}")
    for r, wk in enumerate(sorted(keys), start=2):
        ws.cell(r, 1, wk)
        ws.cell(
            r,
            2,
            f'=AVERAGEIFS({daily_sheet}!$G:$G,{daily_sheet}!$J:$J,A{r})',
        )


def reorder_sheets(wb: Workbook, names: list[str]) -> None:
    sheets = [wb[n] for n in names if n in wb.sheetnames]
    wb._sheets = sheets  # type: ignore[attr-defined]


def main() -> None:
    if not DATA_DIR.is_dir():
        raise SystemExit(f"Missing data directory: {DATA_DIR}")
    if not API_DATA_DIR.is_dir():
        raise SystemExit(f"Missing {API_DATA_DIR}")

    for p in (COLBY_CSV, HILL_CSV):
        if not p.is_file():
            raise SystemExit(f"Missing {p} — run telraam_colby.py / telraam_hillegass_8271.py")

    martin_csv = MARTIN_CSV
    if not martin_csv.is_file():
        legacy = DATA_DIR / "telraam_martin_data.csv"
        if legacy.is_file():
            martin_csv = legacy
        else:
            raise SystemExit(f"Missing {MARTIN_CSV} — run telraam_martin.py")

    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "README"

    colby_keys, colby_idx = embed_api_daily_csv(wb, COLBY_CSV, "Colby_pull", "Colby_data")
    fill_daily_sheet(wb, colby_keys, colby_idx, "Colby_daily", "Colby_data")
    colby_dates = [datetime.strptime(d, "%Y-%m-%d").date() for d in colby_keys]
    fill_weekly(wb, "Colby_weekly", "Colby_daily", colby_dates)

    hill_keys, hill_idx = embed_api_daily_csv(wb, HILL_CSV, "Hillegass_pull", "Hillegass_data")
    fill_daily_sheet(wb, hill_keys, hill_idx, "Hillegass_daily", "Hillegass_data")
    hill_dates = [datetime.strptime(d, "%Y-%m-%d").date() for d in hill_keys]
    fill_weekly(wb, "Hillegass_weekly", "Hillegass_daily", hill_dates)

    martin_keys, martin_idx = embed_api_daily_csv(wb, martin_csv, "Martin_pull", "Martin_data")
    fill_daily_sheet(wb, martin_keys, martin_idx, "Martin_daily", "Martin_data")
    martin_dates = [datetime.strptime(d, "%Y-%m-%d").date() for d in martin_keys]
    fill_weekly(wb, "Martin_weekly", "Martin_daily", martin_dates)

    ws_s = wb.create_sheet("Summary")
    metrics = [
        ("Full_period_ADT_motorized", "AVERAGE", "G"),
        ("Weekday_ADT_motorized", "AVGIFS_WEEKDAY", "G"),
        ("Peak_week_ADT_motorized", "MAX_WEEKLY", "B"),
        ("Busiest_single_day_motorized", "MAX_DAY", "G"),
    ]
    ws_s.cell(1, 1, "Metric")
    ws_s.cell(1, 2, "Colby St")
    ws_s.cell(1, 3, "Hillegass Ave")
    ws_s.cell(1, 4, "Martin St")
    for i, (label, _, _) in enumerate(metrics, start=2):
        ws_s.cell(i, 1, label)

    def avg_motor(daily: str) -> str:
        return (
            f"=AVERAGE(OFFSET({daily}!$G$2,0,0,MAX(0,COUNTA({daily}!$A:$A)-1),1))"
        )

    def avgifs_weekday(daily: str) -> str:
        return f"=AVERAGEIFS({daily}!$G:$G,{daily}!$I:$I,\"<=5\")"

    def peak_week(weekly: str) -> str:
        return f"=MAX({weekly}!$B$2:$B$1000)"

    def busiest_day(daily: str) -> str:
        return f"=MAX(OFFSET({daily}!$G$2,0,0,MAX(0,COUNTA({daily}!$A:$A)-1),1))"

    def busiest_date(daily: str, col: int) -> None:
        ws_s.cell(
            6,
            col,
            f"=TEXT(INDEX({daily}!$A$2:$A$20000,MATCH(MAX({daily}!$G$2:$G$20000),{daily}!$G$2:$G$20000,0)),\"yyyy-mm-dd\")",
        )

    for row, (label, kind, _) in enumerate(metrics, start=2):
        if kind == "AVERAGE":
            ws_s.cell(row, 2, avg_motor("Colby_daily"))
            ws_s.cell(row, 3, avg_motor("Hillegass_daily"))
            ws_s.cell(row, 4, avg_motor("Martin_daily"))
        elif kind == "AVGIFS_WEEKDAY":
            ws_s.cell(row, 2, avgifs_weekday("Colby_daily"))
            ws_s.cell(row, 3, avgifs_weekday("Hillegass_daily"))
            ws_s.cell(row, 4, avgifs_weekday("Martin_daily"))
        elif kind == "MAX_WEEKLY":
            ws_s.cell(row, 2, peak_week("Colby_weekly"))
            ws_s.cell(row, 3, peak_week("Hillegass_weekly"))
            ws_s.cell(row, 4, peak_week("Martin_weekly"))
        elif kind == "MAX_DAY":
            ws_s.cell(row, 2, busiest_day("Colby_daily"))
            ws_s.cell(row, 3, busiest_day("Hillegass_daily"))
            ws_s.cell(row, 4, busiest_day("Martin_daily"))

    ws_s.cell(6, 1, "Busiest day (date)")
    busiest_date("Colby_daily", 2)
    busiest_date("Hillegass_daily", 3)
    busiest_date("Martin_daily", 4)

    ws_s.cell(9, 1, "Weekday_mean_cars")
    ws_s.cell(9, 2, '=AVERAGEIFS(Colby_daily!$B:$B,Colby_daily!$I:$I,"<=5")')
    ws_s.cell(9, 3, '=AVERAGEIFS(Hillegass_daily!$B:$B,Hillegass_daily!$I:$I,"<=5")')
    ws_s.cell(9, 4, '=AVERAGEIFS(Martin_daily!$B:$B,Martin_daily!$I:$I,"<=5")')

    ws_s.cell(10, 1, "Full_period_mean_cars")
    ws_s.cell(10, 2, "=AVERAGE(OFFSET(Colby_daily!$B$2,0,0,MAX(0,COUNTA(Colby_daily!$A:$A)-1),1))")
    ws_s.cell(10, 3, "=AVERAGE(OFFSET(Hillegass_daily!$B$2,0,0,MAX(0,COUNTA(Hillegass_daily!$A:$A)-1),1))")
    ws_s.cell(10, 4, "=AVERAGE(OFFSET(Martin_daily!$B$2,0,0,MAX(0,COUNTA(Martin_daily!$A:$A)-1),1))")

    readme = """\
ADT summaries v5 (generated by telraam_code/build_adt_summaries_xlsx.py)

Data sources (public API pulls, refresh via telraam_code scripts):
  • API_Data/telraam_colby_data.csv
  • API_Data/telraam_hillegass_9000008271_data.csv
  • API_Data/telraam_martin_data.csv

Each CSV is embedded as *_pull / *_data; *_daily formulas aggregate by date.

Motorized = cars + heavy + night (Telraam mode columns).

Refresh: update API CSVs, then run:
  python3 telraam_code/build_adt_summaries_xlsx.py

Weekday = Mon–Fri uses WEEKDAY(...,2) <= 5 (ISO Mon=1 … Fri=5).
"""
    ws_readme["A1"] = readme
    ws_readme["A1"].font = Font(name="Calibri", size=11)
    ws_readme.column_dimensions["A"].width = 92

    head_fill = PatternFill("solid", fgColor="DDEBF7")
    for name in (
        "Summary",
        "Colby_daily",
        "Hillegass_daily",
        "Martin_daily",
        "Colby_pull",
        "Colby_data",
        "Hillegass_pull",
        "Hillegass_data",
        "Martin_pull",
        "Martin_data",
    ):
        if name not in wb.sheetnames:
            continue
        sh = wb[name]
        for c in range(1, min(sh.max_column, 60) + 1):
            cell = sh.cell(1, c)
            if cell.value:
                cell.font = Font(bold=True)
                cell.fill = head_fill

    order = [
        "README",
        "Summary",
        "Colby_pull",
        "Colby_data",
        "Colby_daily",
        "Colby_weekly",
        "Hillegass_pull",
        "Hillegass_data",
        "Hillegass_daily",
        "Hillegass_weekly",
        "Martin_pull",
        "Martin_data",
        "Martin_daily",
        "Martin_weekly",
    ]
    reorder_sheets(wb, order)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
