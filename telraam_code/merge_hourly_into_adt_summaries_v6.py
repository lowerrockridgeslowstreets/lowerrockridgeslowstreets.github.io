"""
Merge telraam_*_hourly.csv rows into adt_summaries_v6.xlsx (in-place).

- Colby_data / Hillegass_data: upsert by local hour in column D (Date and Time (Local)).
  Updates mode totals, V85, uptime, and 0–70 km/h speed % columns when histogram has 8 bins.
- Martin_data: replaces all data rows with the hourly CSV (Martin was daily-only).
  Sets column D header to Date and Time (Local) and fixes Martin_daily SUMIFS to
  calendar-day windows (same pattern as Colby/Hillegass).

Then extends *_daily for new calendar dates through the max date in each hourly CSV,
and appends missing ISO-style week keys to *_weekly (AVERAGEIFS vs daily column J).

Usage (after hourly CSVs exist):
    ~/pipx/shared/bin/python3 telraam_code/merge_hourly_into_adt_summaries_v6.py
"""

from __future__ import annotations

import csv
import json
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

_script_dir = Path(__file__).resolve().parent
if str(_script_dir) not in sys.path:
    sys.path.insert(0, str(_script_dir))

from openpyxl import load_workbook

from telraam_paths import get_telraam_data_dir

DATA_DIR = get_telraam_data_dir()
API_DATA_DIR = DATA_DIR / "API_Data"
XLSX_PATH = DATA_DIR / "adt_summaries_v6.xlsx"

SPEED_70_HEADERS = [
    "Speed Car 0-10 km/h (%)",
    "Speed Car 10-20 km/h (%)",
    "Speed Car 20-30 km/h (%)",
    "Speed Car 30-40 km/h (%)",
    "Speed Car 40-50 km/h (%)",
    "Speed Car 50-60 km/h (%)",
    "Speed Car 60-70 km/h (%)",
    "Speed Car 70+ km/h (%)",
]


def _norm_dt_key(val: Any) -> str:
    if val is None or val == "":
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime.combine(val, datetime.min.time()).strftime("%Y-%m-%d %H:%M:%S")
    s = str(val).strip()
    if len(s) >= 19:
        s = s[:19].replace("T", " ")
    return s


def _header_col_map(ws: Any) -> dict[str, int]:
    m: dict[str, int] = {}
    for c in range(1, 80):
        v = ws.cell(1, c).value
        if v is None:
            continue
        m[str(v).strip()] = c
    return m


def _parse_hist_json(s: str) -> list[float]:
    if not s or not str(s).strip():
        return []
    try:
        arr = json.loads(s)
        if not isinstance(arr, list):
            return []
        out: list[float] = []
        for x in arr:
            try:
                out.append(float(x))
            except (TypeError, ValueError):
                out.append(0.0)
        return out
    except json.JSONDecodeError:
        return []


def _excel_week_key(d: date) -> str:
    """Match daily column J: TEXT(A,\"yyyy\") & \"-W\" & TEXT(WEEKNUM(A,21),\"00\")."""
    return f"{d.year}-W{d.isocalendar()[1]:02d}"


def _merge_rich_data_sheet(
    ws: Any,
    csv_path: Path,
) -> date | None:
    """Colby_data / Hillegass_data upsert. Returns max calendar date or None."""
    if not csv_path.exists():
        print(f"  (skip) missing hourly CSV: {csv_path}")
        return None

    hm = _header_col_map(ws)
    dcol = hm.get("Date and Time (Local)")
    if not dcol:
        raise SystemExit(f"{ws.title}: missing 'Date and Time (Local)' header")

    key_to_row: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        k = _norm_dt_key(ws.cell(r, dcol).value)
        if k:
            key_to_row[k] = r

    with csv_path.open(newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    max_d: date | None = None
    for rec in rows:
        if not rec:
            continue
        k = _norm_dt_key(rec.get("datetime_local", ""))
        if not k:
            continue
        dt = datetime.strptime(k, "%Y-%m-%d %H:%M:%S")
        dday = dt.date()
        if max_d is None or dday > max_d:
            max_d = dday

        if k in key_to_row:
            r = key_to_row[k]
        else:
            r = ws.max_row + 1
            key_to_row[k] = r

        try:
            inst = int(float(rec.get("installation_id") or 0))
        except (TypeError, ValueError):
            inst = rec.get("installation_id") or ""
        ws.cell(r, hm["Installation ID"], inst)
        ws.cell(r, hm["Street"], rec.get("street", ""))
        ws.cell(r, hm["City"], rec.get("city", ""))
        ws.cell(r, dcol, dt)

        ws.cell(r, hm["Pedestrian Total"], float(rec.get("pedestrian") or 0))
        ws.cell(r, hm["Bike Total"], float(rec.get("bike") or 0))
        ws.cell(r, hm["Car Total"], float(rec.get("car") or 0))
        ws.cell(r, hm["Large vehicle Total"], float(rec.get("heavy") or 0))
        ws.cell(r, hm["Night Total"], float(rec.get("night") or 0))

        v85 = rec.get("v85")
        if v85 not in (None, ""):
            try:
                ws.cell(r, hm["Speed V85 km/h"], float(v85))
            except (TypeError, ValueError):
                ws.cell(r, hm["Speed V85 km/h"], v85)
        else:
            ws.cell(r, hm["Speed V85 km/h"], None)

        ut = hm.get("Uptime")
        if ut:
            try:
                ws.cell(r, ut, float(rec.get("uptime") or 0))
            except (TypeError, ValueError):
                ws.cell(r, ut, rec.get("uptime"))

        h70 = _parse_hist_json(rec.get("car_speed_hist_0to70plus") or "")
        if len(h70) == len(SPEED_70_HEADERS):
            for i, name in enumerate(SPEED_70_HEADERS):
                col = hm.get(name)
                if col:
                    ws.cell(r, col, h70[i])

    print(f"  merged {len(rows)} CSV rows into {ws.title} (max date {max_d})")
    return max_d


def _rewrite_martin_hourly(wb: Any, csv_path: Path) -> date | None:
    if not csv_path.exists():
        print(f"  (skip) missing hourly CSV: {csv_path}")
        return None

    with csv_path.open(newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    if not rows:
        print(f"  (skip) empty hourly CSV: {csv_path}")
        return None

    ws = wb["Martin_data"]
    hm = _header_col_map(ws)
    dcol = hm.get("Date") or hm.get("Date and Time (Local)")
    if not dcol:
        raise SystemExit("Martin_data: missing date header in row 1")

    ws.cell(1, dcol, "Date and Time (Local)")

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    rows.sort(key=lambda rec: rec.get("datetime_local") or "")

    max_d: date | None = None
    r = 1
    for rec in rows:
        k = _norm_dt_key(rec.get("datetime_local", ""))
        if not k:
            continue
        r += 1
        dt = datetime.strptime(k, "%Y-%m-%d %H:%M:%S")
        dday = dt.date()
        if max_d is None or dday > max_d:
            max_d = dday

        try:
            inst = int(float(rec.get("installation_id") or 0))
        except (TypeError, ValueError):
            inst = rec.get("installation_id") or ""

        ped = float(rec.get("pedestrian") or 0)
        bike = float(rec.get("bike") or 0)
        car = float(rec.get("car") or 0)
        heavy = float(rec.get("heavy") or 0)
        night = float(rec.get("night") or 0)
        mot = car + heavy + night
        allm = mot + ped + bike

        ws.cell(r, hm["Installation ID"], inst)
        ws.cell(r, hm["Street"], rec.get("street", ""))
        ws.cell(r, hm["City"], rec.get("city", ""))
        ws.cell(r, dcol, dt)
        ws.cell(r, hm["Pedestrian Total"], ped)
        ws.cell(r, hm["Bike Total"], bike)
        ws.cell(r, hm["Car Total"], car)
        ws.cell(r, hm["Large vehicle Total"], heavy)
        ws.cell(r, hm["Night Total"], night)

        v85 = rec.get("v85")
        if v85 not in (None, ""):
            try:
                ws.cell(r, hm["Speed V85 km/h"], float(v85))
            except (TypeError, ValueError):
                ws.cell(r, hm["Speed V85 km/h"], v85)
        else:
            ws.cell(r, hm["Speed V85 km/h"], None)

        ut = hm.get("Uptime")
        if ut:
            try:
                ws.cell(r, ut, float(rec.get("uptime") or 0))
            except (TypeError, ValueError):
                ws.cell(r, ut, rec.get("uptime"))

        ws.cell(r, hm["Motorized Total"], round(mot, 6))
        ws.cell(r, hm["All Modes Total"], round(allm, 6))

    print(f"  rewrote Martin_data with {len(rows)} hourly rows (max date {max_d})")
    return max_d


def _fix_martin_daily_window_formulas(wb: Any) -> None:
    ws = wb["Martin_daily"]
    for r in range(2, ws.max_row + 1):
        a = f"A{r}"
        ws.cell(
            r,
            2,
            f'=SUMIFS(Martin_data!$G:$G,Martin_data!$D:$D,">="&{a},Martin_data!$D:$D,"<"&{a}+1)',
        )
        ws.cell(
            r,
            3,
            f'=SUMIFS(Martin_data!$H:$H,Martin_data!$D:$D,">="&{a},Martin_data!$D:$D,"<"&{a}+1)',
        )
        ws.cell(
            r,
            4,
            f'=SUMIFS(Martin_data!$I:$I,Martin_data!$D:$D,">="&{a},Martin_data!$D:$D,"<"&{a}+1)',
        )
        ws.cell(
            r,
            5,
            f'=SUMIFS(Martin_data!$E:$E,Martin_data!$D:$D,">="&{a},Martin_data!$D:$D,"<"&{a}+1)',
        )
        ws.cell(
            r,
            6,
            f'=SUMIFS(Martin_data!$F:$F,Martin_data!$D:$D,">="&{a},Martin_data!$D:$D,"<"&{a}+1)',
        )
        ws.cell(r, 7, f"=B{r}+C{r}+D{r}")
        ws.cell(r, 8, f"=G{r}+E{r}+F{r}")
        ws.cell(r, 9, f"=WEEKDAY(A{r},2)")
        ws.cell(
            r,
            10,
            f'=TEXT(A{r},"yyyy")&"-W"&TEXT(WEEKNUM(A{r},21),"00")',
        )


def _extend_daily(wb: Any, site: str, end: date | None) -> None:
    if end is None:
        return
    ws = wb[f"{site}_daily"]
    last_v = ws.cell(ws.max_row, 1).value
    if isinstance(last_v, datetime):
        last_d = last_v.date()
    elif isinstance(last_v, date):
        last_d = last_v
    else:
        return
    mp = f"{site}_data"
    d0 = last_d + timedelta(days=1)
    if d0 > end:
        return
    r = ws.max_row
    d = d0
    while d <= end:
        r += 1
        ws.cell(r, 1, datetime.combine(d, datetime.min.time()))
        a = f"A{r}"
        ws.cell(
            r,
            2,
            f'=SUMIFS({mp}!$G:$G,{mp}!$D:$D,">="&{a},{mp}!$D:$D,"<"&{a}+1)',
        )
        ws.cell(
            r,
            3,
            f'=SUMIFS({mp}!$H:$H,{mp}!$D:$D,">="&{a},{mp}!$D:$D,"<"&{a}+1)',
        )
        ws.cell(
            r,
            4,
            f'=SUMIFS({mp}!$I:$I,{mp}!$D:$D,">="&{a},{mp}!$D:$D,"<"&{a}+1)',
        )
        ws.cell(
            r,
            5,
            f'=SUMIFS({mp}!$E:$E,{mp}!$D:$D,">="&{a},{mp}!$D:$D,"<"&{a}+1)',
        )
        ws.cell(
            r,
            6,
            f'=SUMIFS({mp}!$F:$F,{mp}!$D:$D,">="&{a},{mp}!$D:$D,"<"&{a}+1)',
        )
        ws.cell(r, 7, f"=B{r}+C{r}+D{r}")
        ws.cell(r, 8, f"=G{r}+E{r}+F{r}")
        ws.cell(r, 9, f"=WEEKDAY(A{r},2)")
        ws.cell(
            r,
            10,
            f'=TEXT(A{r},"yyyy")&"-W"&TEXT(WEEKNUM(A{r},21),"00")',
        )
        d += timedelta(days=1)
    print(f"  extended {site}_daily through {end} (now row {ws.max_row})")


def _extend_weekly(wb: Any, site: str) -> None:
    dws = wb[f"{site}_daily"]
    wws = wb[f"{site}_weekly"]
    existing: set[str] = set()
    for r in range(2, wws.max_row + 1):
        v = wws.cell(r, 1).value
        if v:
            existing.add(str(v).strip())

    keys: set[str] = set()
    for r in range(2, dws.max_row + 1):
        av = dws.cell(r, 1).value
        if isinstance(av, datetime):
            dd = av.date()
        elif isinstance(av, date):
            dd = av
        else:
            continue
        keys.add(_excel_week_key(dd))

    def _wk_key(s: str) -> tuple[int, int]:
        y, w = s.split("-W", 1)
        return int(y), int(w)

    missing = sorted(keys - existing, key=_wk_key)
    if not missing:
        print(f"  {site}_weekly: no new week keys")
        return
    rr = wws.max_row
    for wk in missing:
        rr += 1
        wws.cell(rr, 1, wk)
        wws.cell(
            rr,
            2,
            f"=AVERAGEIFS({site}_daily!$G:$G,{site}_daily!$J:$J,A{rr})",
        )
    print(f"  extended {site}_weekly with {len(missing)} week rows")


def main() -> None:
    colby_csv = API_DATA_DIR / "telraam_colby_hourly.csv"
    hill_csv = API_DATA_DIR / "telraam_hillegass_9000008271_hourly.csv"
    martin_csv = API_DATA_DIR / "telraam_martin_hourly.csv"

    for p in (colby_csv, hill_csv, martin_csv):
        if not p.exists():
            print(f"Warning: expected hourly CSV not found: {p}")

    if not XLSX_PATH.exists():
        raise SystemExit(f"Workbook not found: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH)

    print("Merging hourly into data sheets…")
    mx_colby = _merge_rich_data_sheet(wb["Colby_data"], colby_csv)
    mx_hill = _merge_rich_data_sheet(wb["Hillegass_data"], hill_csv)
    mx_martin = _rewrite_martin_hourly(wb, martin_csv)

    print("Fixing Martin_daily SUMIFS (calendar-day windows)…")
    _fix_martin_daily_window_formulas(wb)

    print("Extending daily / weekly…")
    for site, mx in (
        ("Colby", mx_colby),
        ("Hillegass", mx_hill),
        ("Martin", mx_martin),
    ):
        _extend_daily(wb, site, mx)
        _extend_weekly(wb, site)

    wb.save(XLSX_PATH)
    print(f"Saved {XLSX_PATH}")


if __name__ == "__main__":
    main()
